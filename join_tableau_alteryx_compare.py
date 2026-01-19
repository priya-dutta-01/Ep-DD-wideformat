import re
from pathlib import Path
from typing import Dict, Tuple, List, Optional, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================================================
# CONFIG – EDIT IF PATHS CHANGE
# ============================================================

SCENARIO_CSV = Path(
    r"C:\Users\57948\OneDrive - Bain\Documents\Ep-DD-wideformat\exports\scenario_wide_exploded.csv"
)

TABLEAU_XLSX = Path(
    r"C:\Users\57948\OneDrive - Bain\Documents\Ep-DD-wideformat\exports\Episode_deep_dive.xlsx"
)

ALTERYX_CSV = Path(
    r"C:\Users\57948\OneDrive - Bain\Documents\Ep-DD-wideformat\exports\Episode_deep_dive_wideformat.csv"
)

MAPPING_XLSX = Path(
    r"C:\Users\57948\OneDrive - Bain\Documents\Ep-DD-wideformat\US_B2C_mapping_file.xlsx"
)

MAPPING_SHEET = "dashboard_sheet_column_mapping"

OUTPUT_XLSX = Path(
    r"C:\Users\57948\OneDrive - Bain\Documents\Ep-DD-wideformat\comparison_output.xlsx"
)

DETAIL_ROW_CAP = 100000


# ============================================================
# Helpers
# ============================================================

def _clean(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def _std_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [_clean(c) for c in df.columns]
    return df


def _safe_sheet(name: str, suffix: str) -> str:
    base = re.sub(r"[\[\]\*\?/\\:]", "_", str(name)).strip()
    max_base = 31 - len(suffix)
    if max_base < 1:
        return suffix[:31]
    return f"{base[:max_base]}{suffix}"


def _safe_flag_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]+", "_", str(s)).strip("_")


def _clean_and_round_series(s: pd.Series, dp: int) -> pd.Series:
    """
    Cleans numeric-like strings before rounding:
      - removes '%' if present
      - removes commas
      - coerces to numeric
      - rounds to dp
    Works even if some rows are plain numerics and some are strings like "12.3%".
    """
    if s is None:
        return s

    s2 = s.astype(str)

    s2 = (
        s2.str.replace("%", "", regex=False)
          .str.replace(",", "", regex=False)
          .str.strip()
    )

    num = pd.to_numeric(s2, errors="coerce")
    return num.round(dp)


def _try_numeric_for_sort(s: pd.Series) -> pd.Series:
    """
    For sorting: if a column looks numeric-ish, sort numerically; else lexicographically.
    Keeps NaN at the end by default behavior in pandas.
    """
    if s is None:
        return s
    coerced = pd.to_numeric(s, errors="coerce")
    if coerced.notna().any():
        return coerced
    return s.astype(str)


def _blank_series(s: pd.Series) -> pd.Series:
    return s.isna() | (s.astype(str).str.strip() == "")


# ============================================================
# Mapping
# ============================================================

def _load_mapping() -> pd.DataFrame:
    m = pd.read_excel(MAPPING_XLSX, sheet_name=MAPPING_SHEET, dtype=str)
    m = _std_cols(m)

    required = [
        "sheet name",
        "column name",
        "alteryx_column_name",
        "is_join_key",
        "compare",
        "include_in_output",
        "round_dp",
    ]
    missing = [c for c in required if c not in m.columns]
    if missing:
        raise ValueError(f"Mapping missing columns: {missing}")

    for c in ["sheet name", "column name", "alteryx_column_name"]:
        m[c] = m[c].map(_clean)

    for c in ["is_join_key", "compare", "include_in_output"]:
        m[c] = (
            m[c]
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"YES": "Y", "TRUE": "Y", "NO": "N", "FALSE": "N", "NAN": ""})
        )

    m["round_dp"] = m["round_dp"].astype(str).replace({"nan": ""}).map(str.strip)

    m = m[
        (m["sheet name"] != "")
        & (m["column name"] != "")
        & (m["alteryx_column_name"] != "")
    ].copy()

    return m


def _pairs_for_sheet(mtab: pd.DataFrame):
    join_pairs: List[Tuple[str, str]] = []
    cmp_pairs: List[Tuple[str, str]] = []
    keep_pairs: List[Tuple[str, str]] = []
    round_map: Dict[Tuple[str, str], int] = {}

    for _, r in mtab.iterrows():
        pair = (r["column name"], r["alteryx_column_name"])

        if r["is_join_key"] == "Y":
            join_pairs.append(pair)
        if r["compare"] == "Y":
            cmp_pairs.append(pair)
        if (r["is_join_key"] == "Y") or (r["compare"] == "Y") or (r.get("include_in_output", "") == "Y"):
            keep_pairs.append(pair)

        dp = str(r.get("round_dp", "")).strip()
        if dp and dp.lower() != "nan":
            try:
                round_map[pair] = int(float(dp))
            except Exception:
                pass

    def dedup(xs):
        out, seen = [], set()
        for x in xs:
            if x not in seen:
                out.append(x)
                seen.add(x)
        return out

    join_pairs = dedup(join_pairs)
    cmp_pairs = [p for p in dedup(cmp_pairs) if p not in join_pairs]
    keep_pairs = dedup(keep_pairs)

    if not join_pairs:
        raise ValueError("No join keys (is_join_key=Y)")
    if not cmp_pairs:
        raise ValueError("No compare fields (compare=Y)")

    return join_pairs, cmp_pairs, keep_pairs, round_map


def _get_alteryx_col_for_tableau_field(mapping: pd.DataFrame, tableau_field: str) -> Optional[str]:
    tf = tableau_field.strip().lower()
    m2 = mapping.copy()
    m2["__col_lc"] = m2["column name"].astype(str).str.strip().str.lower()
    hits = m2[m2["__col_lc"] == tf]
    if hits.empty:
        return None
    return str(hits["alteryx_column_name"].dropna().iloc[0]).strip()


# ============================================================
# Scenario join to populate filter_selection (Alteryx)
# ============================================================

def _build_filter_selection_from_scenarios(scen: pd.DataFrame) -> pd.DataFrame:
    scen = _std_cols(scen)

    required = {"provider", "scenario"}
    if not required.issubset(set(scen.columns)):
        raise ValueError(f"scenario_wide_exploded.csv must contain columns: {required}")

    meta_cols = {
        "scenario_id", "dashboard_name", "provider", "scenario",
        "reference_sheet", "rerolls_used"
    }
    filter_cols = [c for c in scen.columns if c not in meta_cols]

    def agg_unique(series: pd.Series) -> str:
        out = []
        seen = set()
        for v in series.tolist():
            if pd.isna(v):
                continue
            s = str(v).strip()
            if not s:
                continue
            if s not in seen:
                seen.add(s)
                out.append(s)
        return "|".join(out)

    rows = []
    for (prov, scen_num), g in scen.groupby(["provider", "scenario"], dropna=False):
        parts = []
        for c in filter_cols:
            joined = agg_unique(g[c])
            if joined:
                parts.append(f"{c}=[{joined}]")
        rows.append({
            "scen_provider": prov,
            "scen_scenario": scen_num,
            "filter_selection_from_scen": "; ".join(parts),
        })

    return pd.DataFrame(rows)


def _attach_filter_selection_from_scenarios(
    alteryx_df: pd.DataFrame,
    provider_alteryx_col: str,
    scenario_alteryx_col: str,
) -> pd.DataFrame:
    scen_raw = pd.read_csv(SCENARIO_CSV)
    scen_key = _build_filter_selection_from_scenarios(scen_raw)

    alteryx_df = _std_cols(alteryx_df)

    if provider_alteryx_col not in alteryx_df.columns:
        raise ValueError(f"Alteryx provider column not found: '{provider_alteryx_col}'")
    if scenario_alteryx_col not in alteryx_df.columns:
        raise ValueError(f"Alteryx scenario column not found: '{scenario_alteryx_col}'")

    out = alteryx_df.merge(
        scen_key,
        left_on=[provider_alteryx_col, scenario_alteryx_col],
        right_on=["scen_provider", "scen_scenario"],
        how="left",
    )

    if "filter_selection" not in out.columns:
        out["filter_selection"] = pd.NA

    mask = _blank_series(out["filter_selection"])
    out.loc[mask, "filter_selection"] = out.loc[mask, "filter_selection_from_scen"]

    out = out.drop(columns=["scen_provider", "scen_scenario", "filter_selection_from_scen"], errors="ignore")
    return out


# ============================================================
# Output formatting (match highlighting)
# ============================================================

def format_match_cells_true_false(xlsx_path: Path):
    """
    In every *__detail sheet:
      - TRUE cells in __match columns => green
      - FALSE cells in __match columns => red
    """
    wb = load_workbook(xlsx_path)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for ws in wb.worksheets:
        if not ws.title.endswith("__detail"):
            continue

        # Find __match columns by header
        match_col_idxs = []
        for c_idx, cell in enumerate(ws[1], start=1):
            if isinstance(cell.value, str) and cell.value.endswith("__match"):
                match_col_idxs.append(c_idx)

        if not match_col_idxs:
            continue

        # Apply fill row-by-row
        for row in ws.iter_rows(min_row=2):
            for c_idx in match_col_idxs:
                cell = row[c_idx - 1]
                v = cell.value
                if v is True or (isinstance(v, str) and v.strip().upper() == "TRUE"):
                    cell.fill = green_fill
                elif v is False or (isinstance(v, str) and v.strip().upper() == "FALSE"):
                    cell.fill = red_fill

    wb.save(xlsx_path)


# ============================================================
# Suffix enforcement: all Tableau -> _tab, all Alteryx -> _alt
# ============================================================

def _rename_first_existing(df: pd.DataFrame, candidates: List[str], new_name: str) -> None:
    """
    If any candidate exists, rename the first one found to new_name (unless already new_name).
    """
    for c in candidates:
        if c in df.columns:
            if c != new_name:
                df.rename(columns={c: new_name}, inplace=True)
            return


def _enforce_tab_alt_suffixes(
    merged: pd.DataFrame,
    t_cols: Set[str],
    a_cols: Set[str],
) -> pd.DataFrame:
    """
    After merge:
      - ensure each Tableau column ends with _tab
      - ensure each Alteryx column ends with _alt
    Works even when pandas doesn't suffix (because names don't overlap).
    """
    df = merged.copy()

    # Tableau side
    for t in sorted(t_cols):
        if t.endswith("_tab") or t.endswith("_alt"):
            continue
        desired = f"{t}_tab"
        cands = [desired, t, f"{t}__tableau", f"{t}__tab"]
        _rename_first_existing(df, cands, desired)

    # Alteryx side
    for a in sorted(a_cols):
        if a.endswith("_alt") or a.endswith("_tab"):
            continue
        desired = f"{a}_alt"
        cands = [desired, a, f"{a}__alteryx", f"{a}__alt", f"{a}__raw"]
        _rename_first_existing(df, cands, desired)

    return df


# ============================================================
# One filter_selection column only
# ============================================================

def _coalesce_filter_selection_one_col(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create ONE filter_selection column:
      - prefer filter_selection_tab when non-blank
      - otherwise use filter_selection_alt
    Then drop the *_tab and *_alt versions.
    """
    out = df.copy()

    tab_col = "filter_selection_tab" if "filter_selection_tab" in out.columns else None
    alt_col = "filter_selection_alt" if "filter_selection_alt" in out.columns else None

    if tab_col or alt_col:
        out["filter_selection"] = pd.NA

        if tab_col:
            m_tab = ~_blank_series(out[tab_col])
            out.loc[m_tab, "filter_selection"] = out.loc[m_tab, tab_col]

        if alt_col:
            m_need = _blank_series(out["filter_selection"])
            m_alt = ~_blank_series(out[alt_col])
            out.loc[m_need & m_alt, "filter_selection"] = out.loc[m_need & m_alt, alt_col]

        out = out.drop(columns=[c for c in [tab_col, alt_col] if c], errors="ignore")

    return out


# ============================================================
# Sorting detail output
# ============================================================

def _pick_first_existing(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _sort_detail(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sort order (ASC), highest priority first:
      1) scenario_tab
      2) scenario_alt
      3) provider_tab/provider_alt/provider
      4) CV.SurvBank.143.1.1_tab / _alt / plain
    """
    scen_t_col = _pick_first_existing(df, ["scenario_tab", "scenario__tableau", "scenario_tableau", "scenario"])
    scen_a_col = _pick_first_existing(df, ["scenario_alt", "scenario__alteryx", "scenario_alteryx"])
    provider_col = _pick_first_existing(
        df, ["provider_tab", "provider_alt", "provider",
             "survey_provider_tab", "comparison_provider_tab",
             "survey_provider", "comparison_provider"]
    )
    cv_col = _pick_first_existing(
        df, ["CV.SurvBank.143.1.1_tab", "CV.SurvBank.143.1.1_alt", "CV.SurvBank.143.1.1"]
    )

    sort_cols = []
    if scen_t_col:
        sort_cols.append(scen_t_col)
    if scen_a_col and scen_a_col not in sort_cols:
        sort_cols.append(scen_a_col)
    if provider_col and provider_col not in sort_cols:
        sort_cols.append(provider_col)
    if cv_col and cv_col not in sort_cols:
        sort_cols.append(cv_col)

    if not sort_cols:
        return df

    df2 = df.copy()

    tmp_cols = []
    for i, c in enumerate(sort_cols):
        tmp = f"__sortkey_{i}"
        df2[tmp] = _try_numeric_for_sort(df2[c])
        tmp_cols.append(tmp)

    df2 = df2.sort_values(by=tmp_cols, ascending=True, kind="mergesort")
    df2 = df2.drop(columns=tmp_cols, errors="ignore")
    return df2.reset_index(drop=True)


# ============================================================
# Detail column ordering
# ============================================================

def _append_if_exists(out: List[str], df: pd.DataFrame, col: str) -> None:
    if col in df.columns and col not in out:
        out.append(col)


def _build_detail_cols(
    merged: pd.DataFrame,
    join_pairs: List[Tuple[str, str]],
    cmp_pairs: List[Tuple[str, str]],
    keep_pairs: List[Tuple[str, str]],
) -> List[str]:
    """
    REQUIRED BEGINNING (exact order):
      scenario_tab, scenario_alt, provider_tab, provider_alt, filter_selection

    Then:
      - join keys
      - keep/include fields (non-compare)
      - compare fields at end (both sides + match flag)
    """
    cols: List[str] = []

    # 0) MUST-BE-FIRST (exact order)
    for c in ["scenario_tab", "scenario_alt", "filter_selection"]:
        _append_if_exists(cols, merged, c)


    # 1) JOIN keys next
    for (tcol, acol) in join_pairs:
        _append_if_exists(cols, merged, f"{tcol}_tab")
        _append_if_exists(cols, merged, f"{acol}_alt")

    # 2) keep/include fields (exclude join + compare pairs)
    join_set = set(join_pairs)
    cmp_set = set(cmp_pairs)

    for (tcol, acol) in keep_pairs:
        if (tcol, acol) in join_set or (tcol, acol) in cmp_set:
            continue
        _append_if_exists(cols, merged, f"{tcol}_tab")
        _append_if_exists(cols, merged, f"{acol}_alt")

    # 3) COMPARE fields last (+ match flags)
    for (tcol, acol) in cmp_pairs:
        _append_if_exists(cols, merged, f"{tcol}_tab")
        _append_if_exists(cols, merged, f"{acol}_alt")
        flag_col = f"{_safe_flag_name(tcol)}__match"
        _append_if_exists(cols, merged, flag_col)

    return cols


# ============================================================
# Main
# ============================================================

def main():
    mapping = _load_mapping()

    # Load Alteryx CSV (do NOT rename headers)
    alteryx_df = _std_cols(pd.read_csv(ALTERYX_CSV))

    # Find which Alteryx columns correspond to Tableau 'provider' and 'scenario'
    provider_alteryx_col = _get_alteryx_col_for_tableau_field(mapping, "provider")
    scenario_alteryx_col = _get_alteryx_col_for_tableau_field(mapping, "scenario")

    if scenario_alteryx_col is None:
        scenario_alteryx_col = "scenario"

    if provider_alteryx_col is None:
        hint = [c for c in alteryx_df.columns if "prov" in c.lower() or "bank" in c.lower()][:20]
        raise ValueError(
            "Could not find Tableau field 'provider' in mapping file (column name = provider). "
            "Add a mapping row with column name='provider' and alteryx_column_name='<your provider col>'. "
            f"Provider-like columns found in Alteryx: {hint}"
        )

    # Attach scenario-derived filter_selection for Alteryx-only rows
    alteryx_df = _attach_filter_selection_from_scenarios(
        alteryx_df,
        provider_alteryx_col=provider_alteryx_col,
        scenario_alteryx_col=scenario_alteryx_col,
    )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet in sorted(mapping["sheet name"].unique()):
            mtab = mapping[mapping["sheet name"] == sheet].copy()

            try:
                join_pairs, cmp_pairs, keep_pairs, round_map = _pairs_for_sheet(mtab)
            except Exception as e:
                pd.DataFrame([{"sheet": sheet, "error": str(e)}]).to_excel(
                    writer, sheet_name=_safe_sheet(sheet, "__CFGERR"), index=False
                )
                continue

            try:
                tab_df = _std_cols(pd.read_excel(TABLEAU_XLSX, sheet_name=sheet))
            except Exception as e:
                pd.DataFrame([{"sheet": sheet, "error": str(e)}]).to_excel(
                    writer, sheet_name=_safe_sheet(sheet, "__READERR"), index=False
                )
                continue

            required_both = set(join_pairs) | set(cmp_pairs)

            missing_t = {t for (t, a) in required_both if t not in tab_df.columns}
            missing_a = {a for (t, a) in required_both if a not in alteryx_df.columns}

            print(f"Sheet '{sheet}': missing tableau cols: {missing_t}, missing alteryx cols: {missing_a}")
            if missing_t or missing_a:
                pd.DataFrame([{
                    "sheet": sheet,
                    "missing_tableau_cols": ", ".join(sorted(missing_t)),
                    "missing_alteryx_cols": ", ".join(sorted(missing_a)),
                }]).to_excel(
                    writer, sheet_name=_safe_sheet(sheet, "__KEYERR"), index=False
                )
                continue

            tab_join = tab_df.copy()
            alt_join = alteryx_df.copy()

            jk_cols = []
            for i, (tcol, acol) in enumerate(join_pairs):
                jk = f"__jk{i}"
                tab_join[jk] = tab_join[tcol]
                alt_join[jk] = alt_join[acol]
                jk_cols.append(jk)

            merged = tab_join.merge(
                alt_join,
                on=jk_cols,
                how="outer",
                suffixes=("_tab", "_alt"),   # required suffix style
                indicator=True,
            )

            # Force explicit _tab / _alt naming for all mapped cols (even if pandas didn't suffix)
            t_cols_needed = {t for (t, _) in (set(join_pairs) | set(cmp_pairs) | set(keep_pairs))}
            a_cols_needed = {a for (_, a) in (set(join_pairs) | set(cmp_pairs) | set(keep_pairs))}
            merged = _enforce_tab_alt_suffixes(merged, t_cols_needed, a_cols_needed)

            # ------------------------------------------------------------
            # Fill missing filter_selection_tab using lookup from scenario/provider
            # then coalesce to ONE column named filter_selection
            # ------------------------------------------------------------
            fs_t_col = "filter_selection_tab" if "filter_selection_tab" in merged.columns else (
                "filter_selection" if "filter_selection" in merged.columns else None
            )
            fs_a_col = "filter_selection_alt" if "filter_selection_alt" in merged.columns else None

            sc_col = _pick_first_existing(merged, ["scenario_tab", "scenario_alt", "scenario"])
            prov_col = _pick_first_existing(merged, ["provider_tab", "provider_alt", "provider"])

            if fs_t_col and sc_col and "filter_selection_tab" in merged.columns:
                src = merged.copy()
                src_nonblank = src[~_blank_series(src["filter_selection_tab"])]

                if prov_col and prov_col in merged.columns:
                    key = [prov_col, sc_col]
                    lut = (
                        src_nonblank.groupby(key)["filter_selection_tab"]
                        .first()
                        .to_dict()
                    )
                    m = _blank_series(merged["filter_selection_tab"])
                    merged.loc[m, "filter_selection_tab"] = merged.loc[m].apply(
                        lambda r: lut.get((r.get(prov_col), r.get(sc_col))), axis=1
                    )
                else:
                    lut = (
                        src_nonblank.groupby(sc_col)["filter_selection_tab"]
                        .first()
                        .to_dict()
                    )
                    m = _blank_series(merged["filter_selection_tab"])
                    merged.loc[m, "filter_selection_tab"] = merged.loc[m, sc_col].map(lut)

                # still blank? fall back to alteryx filter_selection_alt if present
                m2 = _blank_series(merged["filter_selection_tab"])
                if fs_a_col and fs_a_col in merged.columns:
                    merged.loc[m2, "filter_selection_tab"] = merged.loc[m2, fs_a_col]
                elif "filter_selection" in merged.columns:
                    merged.loc[m2, "filter_selection_tab"] = merged.loc[m2, "filter_selection"]

            # ✅ ONE filter_selection column only
            merged = _coalesce_filter_selection_one_col(merged)

            # ---------------- Field compare + match flags ----------------
            for (tcol, acol) in cmp_pairs:
                t_name = f"{tcol}_tab"
                a_name = f"{acol}_alt"

                t_series = merged[t_name] if t_name in merged.columns else pd.Series([None] * len(merged))
                a_series = merged[a_name] if a_name in merged.columns else pd.Series([None] * len(merged))

                dp = round_map.get((tcol, acol))
                if dp is not None:
                    t_cmp = _clean_and_round_series(t_series, dp)
                    a_cmp = _clean_and_round_series(a_series, dp)
                else:
                    t_cmp = t_series
                    a_cmp = a_series

                both_mask = merged["_merge"] == "both"
                eq = (t_cmp.fillna("__NA__") == a_cmp.fillna("__NA__"))
                eq = eq.where(both_mask, other=pd.NA)

                flag_col = f"{_safe_flag_name(tcol)}__match"
                merged[flag_col] = eq

            # ---------------- DETAIL (column order) ----------------
            detail_cols = _build_detail_cols(
                merged=merged,
                join_pairs=join_pairs,
                cmp_pairs=cmp_pairs,
                keep_pairs=keep_pairs,
            )

            detail = merged.loc[:, detail_cols].copy()
            if len(detail) > DETAIL_ROW_CAP:
                detail = detail.head(DETAIL_ROW_CAP)

            detail = _sort_detail(detail)

            detail.to_excel(writer, sheet_name=_safe_sheet(sheet, "__detail"), index=False)


if __name__ == "__main__":
    main()
    format_match_cells_true_false(OUTPUT_XLSX)
